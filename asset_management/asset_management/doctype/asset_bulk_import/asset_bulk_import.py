# Copyright (c) 2026, Shayan and contributors
# For license information, please see license.txt

import io
import json
import os
import re

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cstr, escape_html, flt, getdate
from frappe.utils.file_manager import get_file_path


REQUIRED_COLUMNS = [
    "asset_name",
    "item_name",
    "category",
    "location",
    "company",
    "rfid_tag",
]

OPTIONAL_COLUMNS = ["images"]

ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

COLUMN_LABELS = {
    "asset_name": "Asset Name",
    "item_name": "Item Name",
    "category": "Category",
    "location": "Location",
    "company": "Company",
    "rfid_tag": "RFID Tag",
    "images": "Images",
}

COLUMN_ALIASES = {
    "asset name": "asset_name",
    "asset_name": "asset_name",
    "item name": "item_name",
    "item_name": "item_name",
    "category": "category",
    "asset category": "category",
    "asset_category": "category",
    "location": "location",
    "company": "company",
    "rfid tag": "rfid_tag",
    "rfid_tag": "rfid_tag",
    "rfid": "rfid_tag",
    "images": "images",
    "image": "images",
}


class AssetBulkImport(Document):
    @frappe.whitelist()
    def preview_import(self):
        """Dry-run: parse the file, validate every row, but do NOT create anything."""
        self._reset_results(status="Draft")

        rows, fatal = self._safe_read_rows()
        if fatal:
            self._render_fatal(fatal)
            self.save(ignore_permissions=True)
            frappe.db.commit()
            return {"success": False, "message": fatal}

        errors, warnings, parsed_rows = self._validate_rows(rows)

        self.rows_total = len(rows)
        self.rows_success = max(0, len(rows) - len({e["row"] for e in errors}))
        self.rows_failed = len({e["row"] for e in errors})

        self.preview_data = json.dumps(parsed_rows, ensure_ascii=False)
        self.preview_html = _build_preview_html(parsed_rows, errors, warnings)
        self.errors_html = _build_errors_html(errors, warnings)
        self.successful_records_html = ""
        self.import_log = "\n".join(
            [_("Preview only — no records created.")]
            + [f"Row {e['row']} [{e['field']}]: {e['message']}" for e in errors]
        )

        self.save(ignore_permissions=True)
        frappe.db.commit()
        return {
            "success": True,
            "rows_total": self.rows_total,
            "rows_failed": self.rows_failed,
            "errors": len(errors),
            "warnings": len(warnings),
        }

    @frappe.whitelist()
    def start_import(self):
        if self.status == "Importing":
            frappe.throw(_("This import is already running."))

        if not self.import_file:
            frappe.throw(_("Please upload an Excel or CSV file first."))

        for f in (
            "default_company",
            "default_purchase_date",
            "default_available_for_use_date",
            "default_gross_purchase_amount",
        ):
            if not self.get(f):
                frappe.throw(
                    _("Default field is required: {0}").format(self.meta.get_label(f))
                )

        self._reset_results(status="Importing")
        self.save(ignore_permissions=True)
        frappe.db.commit()

        rows, fatal = self._safe_read_rows()
        if fatal:
            self._render_fatal(fatal)
            self.status = "Failed"
            self.save(ignore_permissions=True)
            frappe.db.commit()
            return {"success": False, "message": fatal}

        errors, warnings, parsed_rows = self._validate_rows(rows)
        self.rows_total = len(rows)

        successes = []
        log_lines = []
        rows_with_errors = {e["row"] for e in errors}

        for parsed in parsed_rows:
            index = parsed["row"]
            if index in rows_with_errors:
                continue
            try:
                result = self._process_row(parsed)
                successes.append(result)
                log_lines.append(
                    _("Row {0}: created Asset '{1}'").format(index, result["asset_name"])
                )
            except Exception as e:
                errors.append({
                    "row": index,
                    "field": "",
                    "message": cstr(e),
                    "type": "error",
                })
                rows_with_errors.add(index)
                log_lines.append(_("Row {0}: ERROR - {1}").format(index, cstr(e)))
                frappe.log_error(
                    title="Asset Bulk Import row error",
                    message=f"Row {index}: {frappe.get_traceback()}",
                )

        self.rows_success = len(successes)
        self.rows_failed = len(rows_with_errors)

        if self.rows_failed == 0 and self.rows_success > 0:
            self.status = "Success"
        elif self.rows_success == 0:
            self.status = "Failed"
        else:
            self.status = "Partial Success"

        self.preview_data = json.dumps(parsed_rows, ensure_ascii=False)
        self.preview_html = _build_preview_html(parsed_rows, errors, warnings, successes=successes)
        self.errors_html = _build_errors_html(errors, warnings)
        self.successful_records_html = _build_success_html(successes)
        self.import_log = "\n".join(log_lines) if log_lines else _("No rows processed.")

        self.save(ignore_permissions=True)
        frappe.db.commit()

        return {
            "success": True,
            "rows_total": self.rows_total,
            "rows_success": self.rows_success,
            "rows_failed": self.rows_failed,
            "status": self.status,
        }

    def _reset_results(self, status):
        self.status = status
        self.rows_total = 0
        self.rows_success = 0
        self.rows_failed = 0
        self.import_log = ""
        self.preview_html = ""
        self.errors_html = ""
        self.successful_records_html = ""
        self.preview_data = ""

    def _safe_read_rows(self):
        try:
            return _read_rows(self.import_file), None
        except frappe.ValidationError as e:
            return [], cstr(e)
        except Exception as e:
            frappe.log_error(
                title="Asset Bulk Import file read error",
                message=frappe.get_traceback(),
            )
            return [], cstr(e)

    def _render_fatal(self, message):
        self.errors_html = (
            "<div class='asset-import-errors'>"
            "<div class='import-error-card error'>"
            f"<div class='msg'><b>{_('File error')}</b>: {escape_html(message)}</div>"
            "</div></div>"
        )
        self.preview_html = ""
        self.successful_records_html = ""
        self.preview_data = ""
        self.import_log = message

    def _validate_rows(self, rows):
        """Return (errors, warnings, parsed_rows). parsed_rows is the list with index."""
        errors = []
        warnings = []
        parsed_rows = []

        seen_asset_names = {}
        seen_rfids = {}

        for i, raw in enumerate(rows, start=2):
            parsed = {
                "row": i,
                "asset_name": cstr(raw.get("asset_name", "")).strip(),
                "item_name": cstr(raw.get("item_name", "")).strip(),
                "category": cstr(raw.get("category", "")).strip(),
                "location": cstr(raw.get("location", "")).strip(),
                "company": cstr(raw.get("company", "")).strip(),
                "rfid_tag": cstr(raw.get("rfid_tag", "")).strip(),
                "images": cstr(raw.get("images", "")).strip(),
            }
            parsed["image_urls"] = _split_image_urls(parsed["images"])
            parsed_rows.append(parsed)

            for col in REQUIRED_COLUMNS:
                if not parsed[col]:
                    errors.append({
                        "row": i,
                        "field": COLUMN_LABELS[col],
                        "message": _("{0} is required").format(COLUMN_LABELS[col]),
                        "type": "error",
                    })

            an = parsed["asset_name"]
            if an:
                if an in seen_asset_names:
                    errors.append({
                        "row": i,
                        "field": "Asset Name",
                        "message": _("Duplicate Asset Name '{0}' (also in row {1})").format(
                            an, seen_asset_names[an]
                        ),
                        "type": "error",
                    })
                else:
                    seen_asset_names[an] = i
                if frappe.db.exists("Asset", {"asset_name": an}):
                    errors.append({
                        "row": i,
                        "field": "Asset Name",
                        "message": _("Asset with name '{0}' already exists").format(an),
                        "type": "error",
                    })

            rf = parsed["rfid_tag"]
            if rf:
                if rf in seen_rfids:
                    errors.append({
                        "row": i,
                        "field": "RFID Tag",
                        "message": _("Duplicate RFID Tag '{0}' (also in row {1})").format(
                            rf, seen_rfids[rf]
                        ),
                        "type": "error",
                    })
                else:
                    seen_rfids[rf] = i
                if frappe.db.exists("Asset", {"rfid_tag": rf}):
                    errors.append({
                        "row": i,
                        "field": "RFID Tag",
                        "message": _("Asset with RFID tag '{0}' already exists").format(rf),
                        "type": "error",
                    })

            if parsed["category"] and not frappe.db.exists("Asset Category", parsed["category"]):
                warnings.append({
                    "row": i,
                    "field": "Category",
                    "message": _("Asset Category '{0}' does not exist — will be created").format(
                        parsed["category"]
                    ),
                    "type": "warning",
                })

            if parsed["location"]:
                leaf_location = parsed["location"].split(" - ")[-1].strip()
                if not frappe.db.exists("Location", leaf_location):
                    warnings.append({
                        "row": i,
                        "field": "Location",
                        "message": _("Location '{0}' does not exist — will be created").format(
                            parsed["location"]
                        ),
                        "type": "warning",
                    })

            if parsed["company"] and not frappe.db.exists("Company", parsed["company"]):
                errors.append({
                    "row": i,
                    "field": "Company",
                    "message": _("Company '{0}' does not exist").format(parsed["company"]),
                    "type": "error",
                })

            if parsed["item_name"]:
                exists = frappe.db.exists(
                    "Item", {"item_name": parsed["item_name"], "is_fixed_asset": 1}
                )
                if not exists:
                    warnings.append({
                        "row": i,
                        "field": "Item Name",
                        "message": _("Item '{0}' does not exist — will be created with auto-generated Item Code").format(
                            parsed["item_name"]
                        ),
                        "type": "warning",
                    })

            for url in parsed["image_urls"]:
                if not (url.startswith("/files/") or url.startswith("/private/files/") or url.startswith("http")):
                    warnings.append({
                        "row": i,
                        "field": "Images",
                        "message": _("Image '{0}' is not a recognised file URL").format(url),
                        "type": "warning",
                    })

        errors.sort(key=lambda x: (x["row"], x["field"]))
        warnings.sort(key=lambda x: (x["row"], x["field"]))
        return errors, warnings, parsed_rows

    def _process_row(self, parsed):
        company = parsed["company"]
        item_group = _ensure_item_group(parsed["category"])
        asset_category = _ensure_asset_category(parsed["category"], item_group, company)
        location = _ensure_location(parsed["location"], company)
        item_code = _ensure_item(parsed["item_name"], item_group, asset_category, company)

        asset = frappe.new_doc("Asset")
        asset.update({
            "asset_name": parsed["asset_name"],
            "item_code": item_code,
            "asset_category": asset_category,
            "location": location,
            "rfid_tag": parsed["rfid_tag"],
            "company": company,
            "purchase_date": getdate(self.default_purchase_date),
            "available_for_use_date": getdate(self.default_available_for_use_date),
            "gross_purchase_amount": flt(self.default_gross_purchase_amount),
            "is_existing_asset": 1,
            "calculate_depreciation": 0,
            "asset_quantity": 1,
        })

        for url in parsed["image_urls"]:
            asset.append("images", {"image": url})

        asset.flags.ignore_permissions = True
        asset.insert()

        return {
            "row": parsed["row"],
            "asset_name": asset.asset_name,
            "asset_id": asset.name,
            "item_code": item_code,
        }


# ---------- File Reading ----------

def _read_rows(file_url):
    file_path = _resolve_file_path(file_url)
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(file_path)
    if ext == ".csv":
        return _read_csv(file_path)
    raise frappe.ValidationError(
        _("Unsupported file type '{0}'. Use .xlsx or .csv").format(ext)
    )


def _resolve_file_path(file_url):
    if not file_url:
        raise frappe.ValidationError(_("No file attached"))
    return get_file_path(file_url)


def _read_xlsx(path):
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    ws = wb.active

    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [_normalize_header(h) for h in row]
            _check_headers(headers)
            continue
        if all(cell is None or cstr(cell).strip() == "" for cell in row):
            continue
        rows.append(_row_to_dict(headers, row))

    wb.close()
    return rows


def _read_csv(path):
    import csv

    with io.open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = []
        rows = []
        for i, row in enumerate(reader):
            if i == 0:
                headers = [_normalize_header(h) for h in row]
                _check_headers(headers)
                continue
            if all(cstr(cell).strip() == "" for cell in row):
                continue
            rows.append(_row_to_dict(headers, row))
        return rows


def _normalize_header(h):
    if h is None:
        return ""
    key = cstr(h).strip().lower()
    return COLUMN_ALIASES.get(key, key)


def _check_headers(headers):
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        pretty = [COLUMN_LABELS[m] for m in missing]
        raise frappe.ValidationError(
            _("Missing required columns: {0}").format(", ".join(pretty))
        )


def _row_to_dict(headers, row):
    out = {}
    for i, key in enumerate(headers):
        if not key:
            continue
        out[key] = row[i] if i < len(row) else None
    return out


def _split_image_urls(raw):
    if not raw:
        return []
    parts = re.split(r"[,|\n;]+", cstr(raw))
    return [p.strip() for p in parts if p.strip()]


# ---------- Auto-create helpers ----------

def _ensure_item_group(category_name):
    name = cstr(category_name).strip()
    if not name:
        frappe.throw(_("Category cannot be empty"))

    if frappe.db.exists("Item Group", name):
        return name

    parent = "All Item Groups" if frappe.db.exists("Item Group", "All Item Groups") else None
    doc = frappe.new_doc("Item Group")
    doc.item_group_name = name
    if parent:
        doc.parent_item_group = parent
        doc.is_group = 0
    doc.flags.ignore_permissions = True
    doc.insert()
    return doc.name


def _get_account_by_type(company, account_type):
    return frappe.db.get_value(
        "Account",
        {
            "company": company,
            "account_type": account_type,
            "is_group": 0,
            "disabled": 0,
        },
        "name",
    )


def _ensure_asset_category(category_name, item_group, company=None):
    name = cstr(category_name).strip()
    if not name:
        frappe.throw(_("Category cannot be empty"))

    company = company or (
        frappe.defaults.get_user_default("Company")
        or frappe.db.get_single_value("Global Defaults", "default_company")
    )

    if not company:
        frappe.throw(_("Default Company is not set"))

    fixed_asset_account = _get_account_by_type(company, "Fixed Asset")

    if not fixed_asset_account:
        frappe.throw(
            _("No Fixed Asset account found for company '{0}'. Please create an Account with Account Type 'Fixed Asset'.")
            .format(company)
        )

    if frappe.db.exists("Asset Category", name):
        doc = frappe.get_doc("Asset Category", name)

        if not doc.get("company"):
            doc.company = company

        existing_row = None
        for row in doc.accounts:
            if row.company_name == company:
                existing_row = row
                break

        if existing_row:
            if not existing_row.fixed_asset_account:
                existing_row.fixed_asset_account = fixed_asset_account
        else:
            doc.append("accounts", {
                "company_name": company,
                "fixed_asset_account": fixed_asset_account,
            })

        doc.flags.ignore_permissions = True
        doc.save()
        return doc.name

    doc = frappe.new_doc("Asset Category")
    doc.asset_category_name = name
    doc.company = company
    doc.enable_cwip_accounting = 0

    doc.append("accounts", {
        "company_name": company,
        "fixed_asset_account": fixed_asset_account,
    })

    doc.flags.ignore_permissions = True
    doc.insert()
    return doc.name


def _ensure_location(location_path, company=None):
    path = cstr(location_path).strip()
    if not path:
        frappe.throw(_("Location cannot be empty"))

    # Support hierarchy via " - ": "Parent - Child" or "A - B - C"
    parts = [p.strip() for p in path.split(" - ") if p.strip()]

    parent_name = None
    for i, part in enumerate(parts):
        is_leaf = (i == len(parts) - 1)
        existing = frappe.db.exists("Location", part)
        if existing:
            if not is_leaf:
                # Ensure intermediate nodes are groups
                doc = frappe.get_doc("Location", part)
                if not doc.is_group:
                    doc.is_group = 1
                    doc.flags.ignore_permissions = True
                    doc.save()
            parent_name = part
        else:
            doc = frappe.new_doc("Location")
            doc.location_name = part
            doc.is_group = 0 if is_leaf else 1
            if parent_name:
                doc.parent_location = parent_name
            if company:
                doc.company = company
            doc.flags.ignore_permissions = True
            doc.insert()
            parent_name = doc.name

    return parent_name


def _ensure_item(item_name, item_group, asset_category, company=None):
    name = cstr(item_name).strip()
    if not name:
        frappe.throw(_("Item Name cannot be empty"))

    existing = frappe.db.get_value(
        "Item",
        {"item_name": name, "is_fixed_asset": 1},
        "name",
    )
    if existing:
        return existing

    code = _make_item_code(name)

    doc = frappe.new_doc("Item")
    doc.item_code = code
    doc.item_name = name
    doc.item_group = item_group
    doc.is_fixed_asset = 1
    doc.is_stock_item = 0
    doc.asset_category = asset_category
    doc.auto_create_assets = 0
    doc.include_item_in_manufacturing = 0
    if company:
        doc.company = company
    doc.flags.ignore_permissions = True
    doc.insert()
    return doc.name


def _make_item_code(item_name):
    base = re.sub(r"\s+", "-", cstr(item_name).strip())
    base = re.sub(r"[^A-Za-z0-9\-_]", "", base)[:120] or "ITEM"

    if not frappe.db.exists("Item", base):
        return base

    suffix = 1
    while True:
        candidate = f"{base}-{suffix}"
        if not frappe.db.exists("Item", candidate):
            return candidate
        suffix += 1


# ---------- HTML Builders ----------

def _build_preview_html(parsed_rows, errors, warnings, successes=None):
    if not parsed_rows:
        return ""

    rows_with_errors = {e["row"] for e in errors}
    rows_with_warnings = {w["row"] for w in warnings}
    successful_row_ids = {s["row"] for s in (successes or [])}

    head = "".join(
        f"<th>{escape_html(label)}</th>"
        for label in ("#", "Status", *[COLUMN_LABELS[c] for c in ALL_COLUMNS])
    )

    body_rows = []
    for r in parsed_rows:
        row_no = r["row"]
        if row_no in successful_row_ids:
            badge = "<span class='abi-badge abi-success'>Imported</span>"
            row_class = "abi-row-success"
        elif row_no in rows_with_errors:
            badge = "<span class='abi-badge abi-error'>Error</span>"
            row_class = "abi-row-error"
        elif row_no in rows_with_warnings:
            badge = "<span class='abi-badge abi-warning'>Will create</span>"
            row_class = "abi-row-warn"
        else:
            badge = "<span class='abi-badge abi-info'>OK</span>"
            row_class = ""

        cells = [f"<td><b>{row_no}</b></td>", f"<td>{badge}</td>"]
        for c in ALL_COLUMNS:
            v = r.get(c, "")
            if c == "images":
                urls = r.get("image_urls", [])
                v = ", ".join(urls)
            cells.append(f"<td>{escape_html(cstr(v))}</td>")

        body_rows.append(f"<tr class='{row_class}'>{''.join(cells)}</tr>")

    return (
        _STYLE
        + "<div class='asset-import-preview'>"
        + f"<div class='abi-summary'>{len(parsed_rows)} row(s) parsed · "
        + f"<span class='abi-text-error'>{len(rows_with_errors)} with errors</span> · "
        + f"<span class='abi-text-warning'>{len(rows_with_warnings)} with warnings</span></div>"
        + "<table class='abi-table'>"
        + f"<thead><tr>{head}</tr></thead>"
        + f"<tbody>{''.join(body_rows)}</tbody>"
        + "</table></div>"
    )


def _build_errors_html(errors, warnings):
    if not errors and not warnings:
        return ""

    parts = [_STYLE, "<div class='asset-import-errors'>"]

    if errors:
        parts.append(f"<div class='abi-section-title abi-text-error'>{len(errors)} Error(s)</div>")
        for e in errors:
            parts.append(_render_error_card(e, level="error"))

    if warnings:
        parts.append(
            f"<div class='abi-section-title abi-text-warning' style='margin-top:12px'>"
            f"{len(warnings)} Warning(s)</div>"
        )
        for w in warnings:
            parts.append(_render_error_card(w, level="warning"))

    parts.append("</div>")
    return "".join(parts)


def _render_error_card(item, level):
    field = escape_html(item.get("field") or "—")
    message = escape_html(item.get("message") or "")
    row = item.get("row")
    return (
        f"<div class='import-error-card {level}'>"
        f"<div class='row-tag'>Row {row}</div>"
        f"<div class='msg'><b>{field}:</b> {message}</div>"
        "</div>"
    )


def _build_success_html(successes):
    if not successes:
        return ""

    items = []
    for s in successes:
        asset_link = f"/app/asset/{s['asset_id']}"
        item_link = f"/app/item/{escape_html(s['item_code'])}"
        items.append(
            "<div class='import-success-card'>"
            f"<span class='abi-badge abi-success'>Row {s['row']}</span>"
            f"<a href='{asset_link}' target='_blank'><b>{escape_html(s['asset_name'])}</b></a>"
            f"<span class='abi-muted'> · Item: </span>"
            f"<a href='{item_link}' target='_blank'>{escape_html(s['item_code'])}</a>"
            "</div>"
        )

    return (
        _STYLE
        + "<div class='asset-import-success-list'>"
        + f"<div class='abi-section-title abi-text-success'>{len(successes)} record(s) imported</div>"
        + "".join(items)
        + "</div>"
    )


_STYLE = """
<style>
.asset-import-preview, .asset-import-errors, .asset-import-success-list { font-size: 13px; }
.abi-summary { margin-bottom: 8px; color: var(--text-muted); }
.abi-section-title { font-weight: 600; margin-bottom: 8px; }
.abi-text-error { color: var(--red-500, #d9534f); }
.abi-text-warning { color: var(--orange-500, #f0a020); }
.abi-text-success { color: var(--green-500, #28a745); }
.abi-muted { color: var(--text-muted, #888); }
.abi-table { width: 100%; border-collapse: collapse; }
.abi-table th, .abi-table td { padding: 6px 8px; border-bottom: 1px solid var(--border-color, #e5e7eb); text-align: left; vertical-align: top; }
.abi-table thead { background: var(--bg-light-gray, #f8f9fa); }
.abi-row-error { background: rgba(217, 83, 79, 0.06); }
.abi-row-warn { background: rgba(240, 160, 32, 0.06); }
.abi-row-success { background: rgba(40, 167, 69, 0.06); }
.abi-badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }
.abi-badge.abi-success { background: #e6f5ea; color: #1f7a3a; }
.abi-badge.abi-error { background: #fbe9e7; color: #b3261e; }
.abi-badge.abi-warning { background: #fff4e0; color: #a65f00; }
.abi-badge.abi-info { background: #e8f1ff; color: #1a4ea0; }
.import-error-card { display: flex; gap: 10px; align-items: flex-start; padding: 8px 12px; border-radius: 6px; margin-bottom: 6px; border-left: 3px solid; }
.import-error-card.error { background: rgba(217, 83, 79, 0.08); border-color: var(--red-500, #d9534f); }
.import-error-card.warning { background: rgba(240, 160, 32, 0.08); border-color: var(--orange-500, #f0a020); }
.import-error-card .row-tag { font-weight: 600; min-width: 60px; color: var(--text-muted, #555); }
.import-error-card .msg { flex: 1; }
.import-success-card { padding: 6px 10px; border-radius: 6px; margin-bottom: 4px; background: rgba(40, 167, 69, 0.06); display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
.import-success-card a { color: var(--blue-600, #1a73e8); text-decoration: none; }
.import-success-card a:hover { text-decoration: underline; }
</style>
"""
